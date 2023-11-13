import openpyxl
from .. import crud, models


def fill_db():
    dump = openpyxl.load_workbook('src/migrations/test_data.xlsx', data_only=True)

    sales_dump = dump['Продажи']
    for row in sales_dump.iter_rows(min_row=2, max_row=101):
        date = row[0].value.date()
        name, surname = row[1].value.split(' ')
        amount = row[2].value
        crud.save_sale(models.Sale(date=date, manager_id=get_manager_id(name, surname), amount=amount))

    loss_dump = dump['Потери']
    for row in loss_dump.iter_rows(min_row=2, max_row=101):
        date = row[0].value.date()
        name, surname = row[1].value.split(' ')
        amount = row[2].value
        crud.save_loss(models.Loss(date=date, manager_id=get_manager_id(name, surname), loss_amount=amount))

    skill_dump = dump['Навыки']
    for row in skill_dump.iter_rows(min_row=2, max_row=101):
        date = row[0].value.date()
        name, surname = row[1].value.split(' ')
        amount = row[2].value
        crud.save_skill(models.Skill(date=date, manager_id=get_manager_id(name, surname), skill_amount=amount))


def get_manager_id(name, surname):
    manager = crud.get_manager_by_name_and_surname(name, surname)
    if manager is None:
        manager = models.Manager(name=name, surname=surname)
        manager = crud.save_manager(manager)
    return manager.id
